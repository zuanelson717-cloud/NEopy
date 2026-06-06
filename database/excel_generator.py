"""
Módulo Excel - Geração de Relatórios em Excel
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
from config.settings import EXCEL_OUTPUT_DIR


class ExcelGenerator:
    """Classe para gerar relatórios em Excel"""
    
    def __init__(self):
        """Inicializa gerador de Excel"""
        self.ensure_output_dir()
    
    def ensure_output_dir(self):
        """Garante que o diretório de saída existe"""
        if not os.path.exists(EXCEL_OUTPUT_DIR):
            os.makedirs(EXCEL_OUTPUT_DIR)
    
    def gerar_relatorio_materiais(self, materiais, movimentacoes=None):
        """Gera relatório de materiais em Excel"""
        filename = f"{EXCEL_OUTPUT_DIR}relatorio_materiais_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Materiais"
        
        # Estilos
        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws['A1'] = "RELATÓRIO DE MATERIAIS DE INFORMÁTICA"
        ws['A1'].font = Font(bold=True, size=14, color="1F4788")
        ws.merge_cells('A1:H1')
        
        # Data
        ws['A2'] = f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws.merge_cells('A2:H2')
        
        # Cabeçalho da tabela
        headers = ['ID', 'Descrição', 'Categoria', 'Número de Série', 'Quantidade Total', 
                   'Quantidade Disponível', 'Valor Unitário', 'Fornecedor']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Dados
        row = 5
        for material in materiais:
            ws.cell(row=row, column=1).value = material['id']
            ws.cell(row=row, column=2).value = material['descricao']
            ws.cell(row=row, column=3).value = material['categoria']
            ws.cell(row=row, column=4).value = material['numero_serie']
            ws.cell(row=row, column=5).value = material['quantidade_total']
            ws.cell(row=row, column=6).value = material['quantidade_disponivel']
            ws.cell(row=row, column=7).value = f"R$ {material['valor_unitario']:.2f}"
            ws.cell(row=row, column=8).value = material['fornecedor']
            
            # Formatar linha
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col == 7:
                    cell.number_format = 'Currency'
            
            row += 1
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 18
        
        # Salvar
        wb.save(filename)
        return filename


excel_generator = ExcelGenerator()
